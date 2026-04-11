#!/usr/bin/env python3
"""Read all sheets from Tru Tien BOSS.xlsm and write one .js file per sheet (global data for tables)."""

import json
import re
import unicodedata
from collections import defaultdict
from datetime import date, datetime, time
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent
XLSM = ROOT / "Tru Tien BOSS.xlsm"
# Ác mộng 10 + Luyện ngục 10: ảnh trong thư mục dự án (không copy từ xlsm)
PHU_BAN10_IMAGES_DIR = ROOT / "phu ban 10"
# Phụ bản 5: ảnh đặt tay, ghép theo tên boss (tên file gần giống tên boss)
PHU_BAN_IMAGES_DIR = ROOT / "phu ban 5"

# URL-safe .js filenames (loaded before table-app.js; sets window.BOSS_TABLE_DATA)
SHEET_FILES = [
    ("Ác mộng 10", "ac-mong-10.js"),
    ("Luyện ngục 10", "luyen-nguc-10.js"),
    ("Phụ bản 5", "phu-ban-5.js"),
]


def json_default(obj):
    if isinstance(obj, (datetime, date, time)):
        return obj.isoformat()
    raise TypeError(f"Object of type {type(obj).__name__} is not JSON serializable")


def slug_header(h: str, used: dict) -> str:
    """Ensure unique string keys for JSON objects."""
    base = str(h).strip() if h is not None else ""
    if not base:
        base = "column"
    key = base
    n = 1
    while key in used:
        n += 1
        key = f"{base}_{n}"
    used[key] = True
    return key


def openpyxl_rgb_to_css(rgb) -> Optional[str]:
    """Convert openpyxl RGB (e.g. FFFF0000) to #rrggbb."""
    if not rgb:
        return None
    s = str(rgb).upper().replace("0X", "")
    if len(s) == 8:
        return "#" + s[2:8].lower()
    if len(s) == 6:
        return "#" + s.lower()
    return None


# Excel columns for the 12 visible fields on Ác mộng 10 (A–K + P)
AC_MONG_COLS: List[tuple] = [
    ("No.", 1),
    ("Index", 2),
    ("Tên season", 3),
    ("Ngày ra mắt", 4),
    ("Ngày tiêu diệt", 5),
    ("Tuần tiêu diệt", 6),
    ("Số ngày", 7),
    ("Tên BOSS (Ác mộng)", 8),
    ("Tên phụ bản", 9),
    ("Kỹ năng", 10),
    ("Độ khó", 11),
    ("Ghi chú", 16),
]

AC_MONG_JS_KEYS: tuple = tuple(k for k, _ in AC_MONG_COLS)


def split_newlines_to_arrays(obj: Dict[str, Any]) -> None:
    """Chuỗi có \\n → array các dòng (giữ dòng rỗng giữa các phần nếu có)."""
    for k, v in list(obj.items()):
        if k.startswith("__"):
            continue
        if isinstance(v, str) and "\n" in v:
            obj[k] = v.split("\n")


def _is_blank_export_value(v: Any) -> bool:
    if v is None:
        return True
    if isinstance(v, str) and not v.strip():
        return True
    if isinstance(v, list):
        return len(v) == 0 or all(_is_blank_export_value(x) for x in v)
    return False


def export_cell_str(v: Any) -> str:
    """Giá trị ô xuất JS: list → nối bằng \\n (khớp ảnh / hiển thị)."""
    if v is None:
        return ""
    if isinstance(v, list):
        return "\n".join(str(x) for x in v)
    return str(v)


def prune_ac_mong_row(obj: Dict[str, Any]) -> Dict[str, Any]:
    """Keep only HTML-visible columns; drop empty values; keep __bossColor / __images."""
    out: Dict[str, Any] = {}
    for k in AC_MONG_JS_KEYS:
        if k not in obj:
            continue
        v = obj[k]
        if _is_blank_export_value(v):
            continue
        out[k] = v
    if obj.get("__bossColor"):
        out["__bossColor"] = obj["__bossColor"]
    imgs = obj.get("__images")
    if imgs:
        out["__images"] = imgs
    return out


def _leading_digits_filename(filename: str) -> str:
    """'123 foo.JPG' → '123'."""
    m = re.match(r"\s*(\d+)", filename)
    return m.group(1) if m else ""


def _ghichu_lines_ordered(obj: Dict[str, Any]) -> List[str]:
    """Các dòng Ghi chú (thứ tự giữ nguyên)."""
    gc_raw = obj.get("Ghi chú")
    if isinstance(gc_raw, list):
        return [str(x).strip() for x in gc_raw if str(x).strip()]
    if isinstance(gc_raw, str) and gc_raw.strip():
        return [x.strip() for x in gc_raw.split("\n") if x.strip()]
    return []


def _boss_lines_ordered(obj: Dict[str, Any]) -> List[str]:
    """Các dòng Tên BOSS (thứ tự) — Excel nhiều dòng / mảng sau export."""
    raw = obj.get("Tên BOSS (Ác mộng)")
    if isinstance(raw, list):
        return [str(x).strip() for x in raw if str(x).strip()]
    if isinstance(raw, str) and raw.strip():
        return [x.strip() for x in raw.split("\n") if x.strip()]
    return []


def _ac_mong_image_ordering_hints(obj: Dict[str, Any], n_images: int) -> List[str]:
    """
    Gợi ý thứ tự ảnh (theo n_images file đã gán):
    - Nếu số dòng Tên BOSS = n_images → ưu tiên Tên BOSS (Ghi chú có thể là chú thích khác,
      không dùng để khớp tên file — ví dụ Index 322).
    - Ngược lại: dòng Ghi chú nếu có; không thì Tên BOSS.
    """
    boss = _boss_lines_ordered(obj)
    gc = _ghichu_lines_ordered(obj)
    if n_images > 1 and len(boss) == n_images:
        return boss
    if gc:
        return gc
    return boss


def order_ac_mong_images_by_hints(obj: Dict[str, Any], paths: List[str]) -> List[str]:
    """
    Sắp lại thứ tự ảnh theo gợi ý (_ac_mong_image_ordering_hints):
    mỗi dòng gợi ý → ảnh khớp điểm cao nhất trong số còn lại.
    Mỗi ảnh chỉ gán một lần; ảnh thừa nối sau, giữ thứ tự gốc.
    """
    if len(paths) <= 1:
        return paths
    lines = _ac_mong_image_ordering_hints(obj, len(paths))
    if not lines:
        return paths
    pool = list(paths)
    ordered: List[str] = []
    for line in lines:
        if not pool:
            break
        best_p = max(
            pool,
            key=lambda p: score_boss_to_filename(
                line.strip(),
                compact_alnum(Path(p).stem),
                fold_vi_ascii(Path(p).stem),
            ),
        )
        ordered.append(best_p)
        pool.remove(best_p)
    rest = [p for p in paths if p not in ordered]
    return ordered + rest


def attach_boss_images_from_folder(data: list, dir_path: Path, max_n: int = 3) -> None:
    """
    Gán __images trong phu ban 10/: tên file '<Index> <gợi ý phụ bản> <gợi ý boss>',
    lọc theo Index rồi chấm điểm khớp Tên phụ bản + Tên BOSS.
    """
    catalog = build_phu_ban_image_catalog(dir_path)
    if not catalog:
        return
    full_pool: List[tuple] = []
    for rel, fc, ff in catalog:
        stem = Path(rel).stem
        if AC_MONG_IMAGE_NOISE_STEM.match(stem):
            continue
        full_pool.append((rel, fc, ff))
    for obj in data:
        boss = export_cell_str(obj.get("Tên BOSS (Ác mộng)")).strip()
        if not boss:
            continue
        idx_raw = obj.get("Index")
        idx_key = str(idx_raw).strip() if idx_raw is not None else ""
        pool = full_pool
        if idx_key:
            filtered = [
                entry
                for entry in full_pool
                if _leading_digits_filename(Path(entry[0]).name) == idx_key
            ]
            # Có Index nhưng không có file trùng số đầu → không ghép (tránh lấy nhầm index khác)
            pool = filtered
        imgs = images_for_ac_mong_row(obj, pool, max_n=max_n)
        if imgs:
            obj["__images"] = order_ac_mong_images_by_hints(obj, imgs)


def fold_vi_ascii(s: str) -> str:
    t = unicodedata.normalize("NFD", str(s))
    t = "".join(c for c in t if unicodedata.category(c) != "Mn")
    t = t.replace("đ", "d").replace("Đ", "d")
    return t.lower()


def compact_alnum(s: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", fold_vi_ascii(s))


def build_phu_ban_image_catalog(dir_path: Path) -> List[tuple]:
    """(relative_path, stem_compact, stem_folded) per image file."""
    if not dir_path.is_dir():
        return []
    out: List[tuple] = []
    for p in sorted(dir_path.iterdir()):
        if not p.is_file():
            continue
        if p.suffix.lower() not in (".jpg", ".jpeg", ".png", ".gif", ".webp"):
            continue
        rel = f"{dir_path.name}/{p.name}"
        out.append((rel, compact_alnum(p.stem), fold_vi_ascii(p.stem)))
    return out


# Ảnh thừa copy từ Excel (image1.JPG, …) — không dùng khi ghép
AC_MONG_IMAGE_NOISE_STEM = re.compile(r"^image\d+$", re.IGNORECASE)


def score_ac_mong_row_image(
    obj: Dict[str, Any], stem_compact: str, stem_fold: str
) -> int:
    """
    Tên file: '<Index> <tên phụ bản (rút gọn)> <tên boss>' — khớp Index + Tên phụ bản + BOSS.
    """
    boss = export_cell_str(obj.get("Tên BOSS (Ác mộng)")).strip()
    pb = export_cell_str(obj.get("Tên phụ bản") or "").strip()
    if not boss:
        return 0
    idx = str(obj.get("Index") or "").strip()
    stem_c = stem_compact
    if idx:
        idx_c = compact_alnum(idx)
        if idx_c and not stem_c.startswith(idx_c):
            return 0
        rest_c = stem_c[len(idx_c) :] if idx_c else stem_c
    else:
        rest_c = re.sub(r"^\d+", "", stem_c)

    tail_txt = re.sub(r"^\s*\d+\s*", "", str(stem_fold)).strip()
    tail_ns = re.sub(r"[^a-z0-9]+", "", fold_vi_ascii(tail_txt))
    boss_c = compact_alnum(boss)
    pb_c = compact_alnum(pb)

    score = 0
    if boss_c and len(boss_c) >= 2:
        if boss_c in tail_ns:
            score += 14_000
        else:
            for tok in re.findall(r"[a-z0-9]{3,}", fold_vi_ascii(boss)):
                if tok in tail_ns:
                    score += min(1200, 90 * len(tok))
    if pb_c and len(pb_c) >= 3:
        if pb_c in tail_ns:
            score += 6000
        else:
            step = 4
            for i in range(0, max(1, len(pb_c) - 3), step):
                chunk = pb_c[i : i + 7]
                if len(chunk) >= 4 and chunk in tail_ns:
                    score += min(800, 110 * len(chunk))
            for tok in re.findall(r"[a-z0-9]{4,}", fold_vi_ascii(pb)):
                if tok in tail_ns:
                    score += min(700, 55 * len(tok))
    if rest_c:
        overlap = 0
        for ch in boss_c:
            if ch in rest_c:
                overlap += 1
        if boss_c and overlap * 2 >= len(boss_c):
            score += 500
    # Ghi chú (mảng / nhiều dòng): trận nhiều boss — tên file có thể theo dòng ghi chú
    gc_raw = obj.get("Ghi chú")
    lines: List[str] = []
    if isinstance(gc_raw, list):
        lines = [str(x).strip() for x in gc_raw if str(x).strip()]
    elif isinstance(gc_raw, str) and gc_raw.strip():
        lines = [x.strip() for x in gc_raw.split("\n") if x.strip()]
    for line in lines:
        lc = compact_alnum(line)
        if len(lc) >= 3 and lc in tail_ns:
            score += min(6500, 180 * len(lc))
        for tok in re.findall(r"[a-z0-9]{3,}", fold_vi_ascii(line)):
            if tok in tail_ns:
                score += min(900, 75 * len(tok))
    score += score_boss_to_filename(boss, stem_compact, stem_fold) // 12
    return score


def score_boss_to_filename(boss: str, stem_compact: str, stem_fold: str) -> int:
    """Điểm khớp tên boss ↔ tên file (không dấu), có cộng thêm khớp từng từ."""
    boss_s = str(boss).strip()
    if not boss_s:
        return 0
    b_f = fold_vi_ascii(boss_s).replace(" ", "")
    b_c = compact_alnum(boss_s)
    ff_ns = re.sub(r"[^a-z0-9]+", "", stem_fold)
    stem_c = re.sub(r"[^a-z0-9]+", "", stem_compact)
    if not b_f:
        return 0
    score = 0
    if b_c and stem_c and b_c == stem_c:
        score = 10_000
    elif b_f and (b_f in ff_ns or ff_ns in b_f):
        score = 5000 + min(len(b_f), len(ff_ns))
    elif b_c and len(b_c) >= 3 and stem_c and (b_c in stem_c or stem_c in b_c):
        score = 2000 + min(len(b_c), len(stem_c))
    else:
        m = 0
        for i in range(min(len(b_f), len(ff_ns))):
            if b_f[i] == ff_ns[i]:
                m += 1
            else:
                break
        if m >= 5:
            score = m * 20
    # Từ ≥3 ký tự trong tên boss (đã fold) xuất hiện trong tên file
    fold_spaced = fold_vi_ascii(boss_s)
    for tok in re.findall(r"[a-z0-9]{3,}", fold_spaced):
        if tok in stem_fold.lower().replace(" ", ""):
            score += min(len(tok) * 30, 800)
    return score


def _filename_leading_digits(rel: str) -> str:
    """Ví dụ 'phu ban 5/11 foo.JPG' → '11'."""
    base = Path(rel).name
    m = re.match(r"\s*(\d+)", base)
    return m.group(1) if m else ""


# Tên file: "{id} {dungeon} boss {1|2|3} - {mô tả}.jpg"
PHU_BAN_STEM_RE = re.compile(
    r"^\s*(\d+)\s+(.+?)\s+boss\s*([123])\s*-\s*(.+)$",
    re.IGNORECASE | re.UNICODE,
)


def parse_phu_ban_stem(rel: str) -> Optional[tuple]:
    """Trả về (rel, digits, dungeon_raw, slot 1..3, stem_compact, stem_fold) hoặc None."""
    stem = Path(rel).stem
    m = PHU_BAN_STEM_RE.match(stem)
    if not m:
        return None
    digits = m.group(1)
    dungeon_raw = m.group(2).strip()
    slot = int(m.group(3))
    fc = compact_alnum(stem)
    ff = fold_vi_ascii(stem)
    return (rel, digits, dungeon_raw, slot, fc, ff)


def images_for_boss(boss: str, catalog: List[tuple], max_n: int = 2) -> List[str]:
    """Tối đa max_n ảnh khác file, sắp theo điểm giảm dần."""
    boss_s = str(boss).strip()
    if not boss_s:
        return []
    scored: List[tuple] = []
    for rel, fc, ff in catalog:
        sc = score_boss_to_filename(boss_s, fc, ff)
        scored.append((sc, rel))
    scored.sort(key=lambda x: (-x[0], x[1]))
    out: List[str] = []
    seen_paths: set = set()
    best_sc = scored[0][0] if scored else 0
    for sc, rel in scored:
        if rel in seen_paths:
            continue
        if len(out) == 1 and max_n > 1:
            if sc < max(120, best_sc * 0.38):
                break
            d0 = _filename_leading_digits(out[0])
            d1 = _filename_leading_digits(rel)
            if d0 and d1 and d0 != d1:
                continue
        seen_paths.add(rel)
        out.append(rel)
        if len(out) >= max_n:
            break
    if not out and scored:
        out.append(scored[0][1])
    return out


# Điểm tối thiểu để coi là khớp (loại bỏ ghép khi score=0 hoặc chỉ nhiễu từ score_boss/12)
AC_MONG_IMAGE_MIN_SCORE = 450


def images_for_ac_mong_row(obj: dict, pool: List[tuple], max_n: int = 3) -> List[str]:
    """Chọn ảnh theo Index + Tên phụ bản + BOSS (phu ban 10/)."""
    boss = export_cell_str(obj.get("Tên BOSS (Ác mộng)")).strip()
    if not boss:
        return []
    if not pool:
        return []
    scored: List[tuple] = []
    for rel, fc, ff in pool:
        sc = score_ac_mong_row_image(obj, fc, ff)
        scored.append((sc, rel))
    scored.sort(key=lambda x: (-x[0], x[1]))
    out: List[str] = []
    seen_paths: set = set()
    best_sc = scored[0][0] if scored else 0
    if best_sc < AC_MONG_IMAGE_MIN_SCORE:
        return []
    for sc, rel in scored:
        if sc < AC_MONG_IMAGE_MIN_SCORE:
            break
        if rel in seen_paths:
            continue
        if len(out) >= 1 and max_n > 1:
            if sc < max(200, best_sc * 0.22):
                break
            d0 = _filename_leading_digits(out[0])
            d1 = _filename_leading_digits(rel)
            if d0 and d1 and d0 != d1:
                continue
        seen_paths.add(rel)
        out.append(rel)
        if len(out) >= max_n:
            break
    return out


def _group_match_score_for_row(obj: Dict[str, Any], entries: List[tuple]) -> float:
    """Điểm khớp một nhóm ảnh (cùng id + dungeon) với một dòng Excel."""
    by_slot: Dict[int, List[tuple]] = defaultdict(list)
    for e in entries:
        by_slot[e[3]].append(e)
    score = 0.0
    for si in (1, 2, 3):
        key = f"Boss {si}"
        boss = obj.get(key)
        if _is_blank_export_value(boss):
            continue
        cand = by_slot.get(si, [])
        if not cand:
            score -= 4000
            continue
        best = max(
            score_boss_to_filename(export_cell_str(boss), e[4], e[5]) for e in cand
        )
        score += float(best)
    dun_fold = fold_vi_ascii(entries[0][2]) if entries else ""
    pb = obj.get("Tên phụ bản") or ""
    for tok in re.findall(r"[a-z0-9]{4,}", fold_vi_ascii(pb)):
        if tok in re.sub(r"[^a-z0-9]+", "", dun_fold):
            score += 180
    return score


def attach_phu_ban_folder_images(rows: list, dir_path: Path) -> None:
    """Gán __bossImages: ưu tiên tên file có boss 1/2/3, nhóm theo map (số + phụ bản)."""
    catalog = build_phu_ban_image_catalog(dir_path)
    if not catalog:
        return

    groups: Dict[tuple, List[tuple]] = defaultdict(list)
    legacy: List[tuple] = []
    for rel, fc, ff in catalog:
        parsed = parse_phu_ban_stem(rel)
        if parsed is None:
            legacy.append((rel, fc, ff))
            continue
        gkey = (parsed[1], compact_alnum(parsed[2]))
        groups[gkey].append(parsed)

    for obj in rows:
        mapping: Dict[str, List[str]] = {}
        if groups:
            best_key: Optional[tuple] = None
            best_sc = -1e20
            for gkey, entries in groups.items():
                sc = _group_match_score_for_row(obj, entries)
                if sc > best_sc:
                    best_sc = sc
                    best_key = gkey
            if best_key is not None and best_sc > -2500:
                by_slot: Dict[int, List[tuple]] = defaultdict(list)
                for e in groups[best_key]:
                    by_slot[e[3]].append(e)
                for key in ("Boss 1", "Boss 2", "Boss 3"):
                    si = int(key[-1])
                    boss = obj.get(key)
                    if _is_blank_export_value(boss):
                        continue
                    pool = [(e[0], e[4], e[5]) for e in by_slot.get(si, [])]
                    if not pool:
                        continue
                    imgs = images_for_boss(export_cell_str(boss), pool, max_n=2)
                    if imgs:
                        mapping[key] = imgs
        if not mapping:
            pool = legacy + catalog if legacy else catalog
            for key in ("Boss 1", "Boss 2", "Boss 3"):
                boss = obj.get(key)
                if _is_blank_export_value(boss):
                    continue
                imgs = images_for_boss(export_cell_str(boss), pool, max_n=2)
                if imgs:
                    mapping[key] = imgs
        if mapping:
            obj["__bossImages"] = mapping


# Luyện ngục 10 — A–J + O (không cột Kỹ năng)
LUYEN_NGUC_COLS: List[tuple] = [
    ("No.", 1),
    ("Index", 2),
    ("Tên season", 3),
    ("Ngày ra mắt", 4),
    ("Ngày tiêu diệt", 5),
    ("Tuần tiêu diệt", 6),
    ("Số ngày", 7),
    ("Tên BOSS (Ác mộng)", 8),
    ("Tên phụ bản", 9),
    ("Độ khó", 10),
    ("Ghi chú", 15),
]
LUYEN_NGUC_JS_KEYS: tuple = tuple(k for k, _ in LUYEN_NGUC_COLS)


def prune_luyen_nguc_row(obj: Dict[str, Any]) -> Dict[str, Any]:
    out: Dict[str, Any] = {}
    for k in LUYEN_NGUC_JS_KEYS:
        if k not in obj:
            continue
        v = obj[k]
        if _is_blank_export_value(v):
            continue
        out[k] = v
    if obj.get("__bossColor"):
        out["__bossColor"] = obj["__bossColor"]
    imgs = obj.get("__images")
    if imgs:
        out["__images"] = imgs
    return out


# Phụ bản 5 — A–F
PHU_BAN_COLS: List[tuple] = [
    ("No.", 1),
    ("Mùa", 2),
    ("Tên phụ bản", 3),
    ("Boss 1", 4),
    ("Boss 2", 5),
    ("Boss 3", 6),
]
PHU_BAN_JS_KEYS: tuple = tuple(k for k, _ in PHU_BAN_COLS)


def prune_phu_ban_row(obj: Dict[str, Any]) -> Dict[str, Any]:
    out: Dict[str, Any] = {}
    for k in PHU_BAN_JS_KEYS:
        if k not in obj:
            continue
        v = obj[k]
        if _is_blank_export_value(v):
            continue
        out[k] = v
    bim = obj.get("__bossImages")
    if bim:
        out["__bossImages"] = bim
    return out


def filter_rows_with_boss_name(sheet_name: str, rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Drop rows with no boss name (sheet-specific)."""
    if sheet_name in ("Ác mộng 10", "Luyện ngục 10"):
        return [
            r
            for r in rows
            if not _is_blank_export_value(r.get("Tên BOSS (Ác mộng)"))
        ]
    if sheet_name == "Phụ bản 5":
        return [
            r
            for r in rows
            if any(
                not _is_blank_export_value(r.get(k))
                for k in ("Boss 1", "Boss 2", "Boss 3")
            )
        ]
    return rows


def enrich_ac_mong_boss_font(ws, rows: list, boss_col: int = 8) -> None:
    """Attach __bossColor from Excel font rgb on 'Tên BOSS' column (column H = 8)."""
    for i, obj in enumerate(rows):
        cell = ws.cell(row=i + 2, column=boss_col)
        co = cell.font.color
        if co and co.type == "rgb" and co.rgb:
            css = openpyxl_rgb_to_css(co.rgb)
            if css:
                obj["__bossColor"] = css


def sheet_to_rows(ws):
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header_row = rows[0]
    used = {}
    keys = [slug_header(c, used) for c in header_row]
    out = []
    for row in rows[1:]:
        obj = {}
        for i, k in enumerate(keys):
            v = row[i] if i < len(row) else None
            obj[k] = v
        out.append(obj)
    return out


def main():
    wb = load_workbook(XLSM, data_only=True)
    wb_fmt = load_workbook(XLSM, data_only=False)
    try:
        for sheet_name, js_name in SHEET_FILES:
            if sheet_name not in wb.sheetnames:
                raise SystemExit(f"Missing sheet {sheet_name!r}. Found: {wb.sheetnames}")
            data = sheet_to_rows(wb[sheet_name])
            for obj in data:
                split_newlines_to_arrays(obj)
            if sheet_name == "Ác mộng 10":
                attach_boss_images_from_folder(data, PHU_BAN10_IMAGES_DIR, max_n=3)
                enrich_ac_mong_boss_font(wb_fmt[sheet_name], data)
                data = [prune_ac_mong_row(row) for row in data]
                data = filter_rows_with_boss_name(sheet_name, data)
            elif sheet_name == "Luyện ngục 10":
                attach_boss_images_from_folder(data, PHU_BAN10_IMAGES_DIR, max_n=3)
                enrich_ac_mong_boss_font(wb_fmt[sheet_name], data)
                data = [prune_luyen_nguc_row(row) for row in data]
                data = filter_rows_with_boss_name(sheet_name, data)
            elif sheet_name == "Phụ bản 5":
                data = [prune_phu_ban_row(row) for row in data]
                data = filter_rows_with_boss_name(sheet_name, data)
                attach_phu_ban_folder_images(data, PHU_BAN_IMAGES_DIR)
            payload = json.dumps(data, ensure_ascii=False, indent=2, default=json_default)
            path = ROOT / js_name
            path.write_text(
                "// Auto-generated from Tru Tien BOSS.xlsm — do not edit by hand.\n"
                f'// Sheet: "{sheet_name}"\n'
                "window.BOSS_TABLE_DATA = "
                + payload
                + ";\n",
                encoding="utf-8",
            )
            print(f"Wrote {path.name} ({len(data)} rows)")
    finally:
        wb.close()
        wb_fmt.close()


if __name__ == "__main__":
    main()
